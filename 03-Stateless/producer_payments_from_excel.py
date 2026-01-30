import socket
from openpyxl import load_workbook

from confluent_kafka import Producer
from confluent_kafka.schema_registry import SchemaRegistryClient
from confluent_kafka.schema_registry.json_schema import JSONSerializer
from confluent_kafka.serialization import SerializationContext, MessageField

TOPIC = "payments.raw"
EXCEL_PATH = "payments_input.xlsx"
SHEET_NAME = "payments"

FIELDS = [
    "event_id", "card_id", "user_id", "merchant_id", "merchant_category",
    "amount", "currency", "country", "ip_country", "channel", "ts_utc"
]

def delivery_report(err, msg):
    if err:
        print(f"Delivery failed: {err}")
    else:
        print(f"Delivered to {msg.topic()} [{msg.partition()}] offset {msg.offset()}")

def row_to_event(row_dict: dict) -> dict:
    # Excel may store numeric cells as ints/floats; keep types consistent
    row_dict["amount"] = float(row_dict["amount"])
    return row_dict

def main():
    schema_str = open("schema_payment.json", "r", encoding="utf-8").read()

    sr = SchemaRegistryClient({"url": "http://localhost:8081"})
    serializer = JSONSerializer(schema_str, sr, lambda obj, ctx: obj)

    producer = Producer({
        "bootstrap.servers": "localhost:9092",
        "client.id": f"payments-producer-excel-{socket.gethostname()}",
        "acks": 1
    })

    wb = load_workbook(EXCEL_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {EXCEL_PATH}. Found: {wb.sheetnames}")

    ws = wb[SHEET_NAME]

    header_row = [cell.value for cell in ws[1]]
    if header_row[:len(FIELDS)] != FIELDS:
        raise ValueError(
            "Header row does not match expected schema.\n"
            f"Expected: {FIELDS}\n"
            f"Found:    {header_row}"
        )

    sent = 0
    # Track Excel row index so we can attach it to message headers for tracing
    for excel_row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        event = dict(zip(FIELDS, row[:len(FIELDS)]))

        missing = [k for k, v in event.items() if v is None]
        if missing:
            print(f"Skipping row {excel_row_idx} missing {missing}: {event}")
            continue

        event = row_to_event(event)

        key = str(event["card_id"]).encode("utf-8")
        value = serializer(event, SerializationContext(TOPIC, MessageField.VALUE))

        producer.produce(
            TOPIC,
            key=key,
            value=value,
            headers=[
                ("source", EXCEL_PATH.encode("utf-8")),
                ("excel_sheet", SHEET_NAME.encode("utf-8")),
                ("excel_row", str(excel_row_idx).encode("utf-8")),
            ],
            on_delivery=delivery_report
        )
        producer.poll(0)
        sent += 1

    producer.flush()
    print(f"Sent {sent} events from {EXCEL_PATH} to {TOPIC}")

if __name__ == "__main__":
    main()
